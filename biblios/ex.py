#@File: ex.py
#@Author: Diego Martínez Sánchez
#@Brief: Methods for open the excel file and get the information we need from the same excel file
try:
    import os
    import time
    import openpyxl
except ImportError as eImp:
    print(f"The following import ERROR ocurred in ex file: {eImp}")

class Excel(object):
    def DatosEx(self, bande, ruta):
        arch= openpyxl.load_workbook(ruta, data_only=True)
        arre= arch.get_sheet_names()
        hoja= arch.get_sheet_by_name(arre[0])
        if bande== 0:
            rutaIni= hoja["E2"].value
            return rutaIni
        elif bande== 1:
            rutaIni= hoja["F2"].value
            return rutaIni
        elif bande== 2:
            rutaIni= hoja["G2"].value
            return rutaIni

    def NomCarpesRuta(self, ruta):
        contador=1
        arre= []
        carpe= ""
        salir= False
        doc= openpyxl.load_workbook(ruta, data_only= True)
        aus1= 2
        os.system("cls")
        arrehojas= doc.get_sheet_names()
        hoja= doc.get_sheet_by_name(arrehojas[0])
        cell1= hoja["A2"].value

        while not salir:
            aus1+= 1
            #s= 0
            aux2= str(aus1)
            NumCell= "A"+aux2
            cell2= hoja[NumCell].value

            if cell2!= cell1:
                aus2= aus1
                aus2-= 1
                NumCarpe= "A"+str(aus2)
                carpe= hoja[NumCarpe].value
                try:
                    auxreq= carpe.split(")")
                    carpe= auxreq[0]
                except Exception as e:
                    print(f"Ocurrió el ERROR: {e}")
                    time.sleep(4)
                carpe+= " - OK"
                arre.append(carpe)
                cell1= cell2
                contador+= 1

                if cell1==None or cell2== None:
                    salir= True
                if cell2== cell1:
                    pass
                elif carpe== None:
                    aus2= aus1
                    aus2-= 1
                    #num= "A"+str(aus2)
                    carpe= hoja[NumCarpe].value
                    carpe+= " - OK"
                    arre.append(carpe)
                    salir= True
                else:
                    pass
        return arre

    def Ex(self, ruta):
        salir= False
        doc= openpyxl.load_workbook(ruta, data_only=True)
        aus1= 1
        os.system("cls")
        arrehojas= doc.get_sheet_names()
        hoja= doc.get_sheet_by_name(arrehojas[0])
        cell1= hoja["A2"].value
        valor= ""
        arre= []

        while not salir:
            aus1+=1
            aux2= str(aus1)
            num= "D"+aux2#Antes E pero se decidio cambiar.
            NumCell= "A"+aux2
            cell2= hoja[NumCell].value
            valor= hoja[num].value
            if cell2!=cell1 and valor!=None:
                aus2= aus1
                aus2-=1
                num= "D"+str(aus2)#Antes E pero se decidio cambiar.
                valor= hoja[num].value
                arre.append(valor)
                print("")
                cell1= cell2
            elif valor==None:
                aus2= aus1
                aus2-=1
                num= "D"+str(aus2)#Antes E pero se decidio cambiar.
                valor= hoja[num].value
                arre.append(valor)
                salir= True
            else:
                pass
        return arre
    
    def TituloWord(self, ruta):
        doctit= openpyxl.load_workbook(ruta, data_only= True)
        arrehojastit= doctit.get_sheet_names()
        hojatit= doctit.get_sheet_by_name(arrehojastit[0])
        celltit= hojatit["I2"].value
        doctit.close()

        return celltit